# -*- coding: utf-8 -*-
"""
Excel Lookup
============
ค้นหาข้อมูลจากไฟล์ Excel (.xlsx) โดยใช้ PO Number เป็น key
ไฟล์ Excel อาจอยู่ที่:
  - บน OneDrive/SharePoint ที่ sync ลงเครื่อง (เช่น C:\\Users\\xxx\\OneDrive\\...)
  - บน network drive (เช่น \\\\server\\share\\file.xlsx)
  - บนเครื่อง local
ไม่ต้อง auth, ไม่ต้องเชื่อม API — เปิดไฟล์ตรงๆ

วิธีตั้งค่า:
  - กรอก file_path, sheet_name, po_column ใน config.ini
"""

import os
import configparser
from datetime import datetime, date

try:
    import openpyxl
except ImportError:
    raise ImportError(
        "ไม่พบโมดูล openpyxl — กรุณาติดตั้ง:\n"
        "  pip install openpyxl"
    )


# -----------------------------------------------------------------------------
# Constants
# -----------------------------------------------------------------------------
CONFIG_FILENAME = "config.ini"
DEFAULT_PO_COLUMN = "poNo"


# -----------------------------------------------------------------------------
# Paths
# -----------------------------------------------------------------------------
def _here():
    return os.path.dirname(os.path.abspath(__file__))


def _config_path():
    return os.path.join(_here(), CONFIG_FILENAME)


# -----------------------------------------------------------------------------
# Configuration
# -----------------------------------------------------------------------------
def load_config():
    """โหลด config.ini และคืนค่าเป็น dict สำหรับ section [excel]"""
    path = _config_path()
    if not os.path.exists(path):
        raise FileNotFoundError(
            f"ไม่พบไฟล์ {CONFIG_FILENAME} ที่ {path}\n"
            f"กรุณาสร้างไฟล์ตามตัวอย่างใน config.ini.template"
        )

    parser = configparser.ConfigParser()
    parser.read(path, encoding="utf-8")

    if "excel" not in parser:
        raise ValueError(f"ไม่พบ section [excel] ใน {CONFIG_FILENAME}")

    cfg = {
        "file_path": parser["excel"].get("file_path", "").strip(),
        "sheet_name": parser["excel"].get("sheet_name", "").strip() or None,
        "po_column": parser["excel"].get("po_column", DEFAULT_PO_COLUMN).strip()
                     or DEFAULT_PO_COLUMN,
    }

    if not cfg["file_path"] or cfg["file_path"].startswith("<"):
        raise ValueError(
            f"กรุณาตั้งค่า file_path ใน [excel] section ของ {CONFIG_FILENAME}"
        )

    # ขยาย ~ / environment variables
    cfg["file_path"] = os.path.expanduser(os.path.expandvars(cfg["file_path"]))
    return cfg


# -----------------------------------------------------------------------------
# Excel helpers
# -----------------------------------------------------------------------------
def _normalize_cell(value):
    """แปลงค่าจากเซลล์ Excel ให้อยู่ในรูปแบบที่เหมาะสำหรับแสดง"""
    if value is None:
        return None
    if isinstance(value, (datetime, date)):
        return value.strftime("%Y-%m-%d") if isinstance(value, date) \
            and not isinstance(value, datetime) else value.strftime("%Y-%m-%d %H:%M") \
            if isinstance(value, datetime) and (value.hour or value.minute) \
            else value.strftime("%Y-%m-%d")
    return value


def _po_matches(cell_value, target):
    """เปรียบเทียบ PO โดยตัดช่องว่างและแปลงเป็น string ก่อน"""
    if cell_value is None:
        return False
    # Excel อาจเก็บ PO เป็น number (4500987654.0) หรือ string
    if isinstance(cell_value, float) and cell_value.is_integer():
        cell_str = str(int(cell_value))
    else:
        cell_str = str(cell_value).strip()
    return cell_str == str(target).strip()


# -----------------------------------------------------------------------------
# Lookup
# -----------------------------------------------------------------------------
def lookup_po(po_number, excel_path=None, sheet_name=None, po_column=None):
    """
    ค้นหา row ของ PO ในไฟล์ Excel
    Args:
        po_number:    PO number ที่ต้องการค้นหา
        excel_path:   path ของไฟล์ Excel (ถ้าไม่ระบุจะใช้จาก config.ini)
        sheet_name:   ชื่อ sheet (ถ้าไม่ระบุจะใช้ active sheet)
        po_column:    ชื่อคอลัมน์ที่เก็บ PO (default: 'poNo')
    Returns:
        dict ของ {column_name: value} หรือ None ถ้าไม่พบ
    """
    if not po_number or not str(po_number).strip():
        raise ValueError("PO number ห้ามว่าง")

    # ใช้ค่าจาก config ถ้าไม่ได้ส่งมา
    if excel_path is None or po_column is None:
        cfg = load_config()
        excel_path = excel_path or cfg["file_path"]
        sheet_name = sheet_name if sheet_name is not None else cfg["sheet_name"]
        po_column = po_column or cfg["po_column"]

    if not os.path.exists(excel_path):
        raise FileNotFoundError(
            f"ไม่พบไฟล์ Excel: {excel_path}\n"
            f"กรุณาตรวจสอบ:\n"
            f"  • path ถูกต้อง\n"
            f"  • ถ้าเป็นไฟล์บน OneDrive: ต้อง sync ลงเครื่องแล้ว\n"
            f"  • คลิกขวาที่ไฟล์ → 'Always keep on this device'"
        )

    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True, read_only=True)
    except PermissionError:
        raise PermissionError(
            f"ไม่สามารถเปิดไฟล์ได้ — ไฟล์อาจถูกเปิดอยู่ในโปรแกรมอื่น\n"
            f"กรุณาปิด Excel แล้วลองใหม่: {excel_path}"
        )
    except Exception as exc:
        raise RuntimeError(f"เปิดไฟล์ Excel ไม่ได้: {exc}")

    try:
        # เลือก sheet
        if sheet_name:
            if sheet_name not in wb.sheetnames:
                raise ValueError(
                    f"ไม่พบ sheet '{sheet_name}'\n"
                    f"Sheets ที่มี: {wb.sheetnames}"
                )
            ws = wb[sheet_name]
        else:
            ws = wb.active

        # อ่าน header row (แถวแรก)
        rows_iter = ws.iter_rows(values_only=True)
        try:
            headers = next(rows_iter)
        except StopIteration:
            return None  # ไฟล์ว่าง

        if not headers:
            return None

        # หาตำแหน่งของคอลัมน์ PO
        po_col_idx = None
        for i, h in enumerate(headers):
            if h is not None and str(h).strip() == po_column:
                po_col_idx = i
                break

        if po_col_idx is None:
            available = [str(h).strip() for h in headers if h]
            raise ValueError(
                f"ไม่พบคอลัมน์ '{po_column}' ใน sheet '{ws.title}'\n"
                f"คอลัมน์ที่มี ({len(available)}): {available[:15]}"
                f"{'...' if len(available) > 15 else ''}"
            )

        # ค้นหา row
        for row in rows_iter:
            if po_col_idx >= len(row):
                continue
            if _po_matches(row[po_col_idx], po_number):
                # พบแล้ว สร้าง dict (ข้าม header ที่เป็น None)
                result = {}
                for i in range(min(len(headers), len(row))):
                    if headers[i] is not None:
                        col_name = str(headers[i]).strip()
                        result[col_name] = _normalize_cell(row[i])
                return result

        return None
    finally:
        wb.close()


def list_sheets_and_columns(excel_path=None, sheet_name=None):
    """
    ตรวจสอบโครงสร้างไฟล์ Excel — แสดง sheets และ columns ทั้งหมด
    ใช้สำหรับ debug ก่อนตั้ง config
    """
    if excel_path is None:
        cfg = load_config()
        excel_path = cfg["file_path"]
        sheet_name = sheet_name or cfg["sheet_name"]

    if not os.path.exists(excel_path):
        raise FileNotFoundError(f"ไม่พบไฟล์: {excel_path}")

    wb = openpyxl.load_workbook(excel_path, data_only=True, read_only=True)
    try:
        result = {"file": excel_path, "sheets": {}}
        for s_name in wb.sheetnames:
            ws = wb[s_name]
            try:
                headers = next(ws.iter_rows(values_only=True))
                cols = [str(h).strip() for h in headers if h is not None]
            except StopIteration:
                cols = []
            result["sheets"][s_name] = cols
        return result
    finally:
        wb.close()


# -----------------------------------------------------------------------------
# CLI test
# -----------------------------------------------------------------------------
if __name__ == "__main__":
    import sys

    if len(sys.argv) < 2:
        print("ใช้งาน:")
        print("  python excel_lookup.py <PO_NUMBER>     # ค้นหา PO")
        print("  python excel_lookup.py --list          # ดู sheets/columns")
        sys.exit(1)

    if sys.argv[1] == "--list":
        try:
            info = list_sheets_and_columns()
        except Exception as e:
            print(f"❌ {e}")
            sys.exit(1)
        print(f"ไฟล์: {info['file']}\n")
        for sheet, cols in info["sheets"].items():
            print(f"  Sheet: {sheet} ({len(cols)} columns)")
            for c in cols[:30]:
                print(f"    • {c}")
            if len(cols) > 30:
                print(f"    ... และอีก {len(cols) - 30} columns")
        sys.exit(0)

    po = sys.argv[1]
    print(f"กำลังค้นหา PO: {po}")
    try:
        row = lookup_po(po)
    except Exception as e:
        print(f"❌ {e}")
        sys.exit(1)

    if not row:
        print(f"❌ ไม่พบ PO: {po}")
        sys.exit(0)

    print(f"\n✓ พบข้อมูล PO {po} ({len(row)} columns):\n")
    for key, value in row.items():
        if value not in (None, ""):
            print(f"  {key:35s} = {value}")
