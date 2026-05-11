# -*- coding: utf-8 -*-
"""
ระบบสร้างเอกสารอัตโนมัติ - ปตท. (Web Version)
==============================================
Streamlit web app version ของ auto_form_generator

Run locally:
    pip install -r requirements.txt
    streamlit run streamlit_app.py

Deploy:
    Push to GitHub → connect repo to Streamlit Community Cloud
    https://share.streamlit.io
"""

import os
import sys
from io import BytesIO
from datetime import datetime, date

import streamlit as st
import openpyxl

# เพิ่ม path ให้ import จาก root directory ได้
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from form_logic import (
    generate_form,
    FORM_TYPES,
    DOC_TYPE_OPTIONS,
    DISPLAY_FIELDS,
)


# =============================================================================
# Helper: lookup PO ใน Excel bytes ที่ผู้ใช้ upload หรือใน sample file
# =============================================================================
def lookup_po_in_bytes(po_number, file_bytes, sheet_name=None, po_column="poNo"):
    """ค้นหา PO ในไฟล์ Excel จาก bytes ที่ส่งเข้ามา"""
    wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True, read_only=True)
    try:
        # เลือก sheet
        if sheet_name and sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.active

        rows_iter = ws.iter_rows(values_only=True)
        try:
            headers = next(rows_iter)
        except StopIteration:
            return None, []

        if not headers:
            return None, []

        available_cols = [str(h).strip() for h in headers if h is not None]

        # หาตำแหน่งคอลัมน์ PO
        po_col_idx = None
        for i, h in enumerate(headers):
            if h is not None and str(h).strip() == po_column:
                po_col_idx = i
                break

        if po_col_idx is None:
            return {"__error__": f"ไม่พบคอลัมน์ '{po_column}' ใน sheet '{ws.title}'"}, available_cols

        # ค้นหา row
        po_str = str(po_number).strip()
        for row in rows_iter:
            if po_col_idx >= len(row):
                continue
            cell = row[po_col_idx]
            if cell is None:
                continue
            if isinstance(cell, float) and cell.is_integer():
                cell_str = str(int(cell))
            else:
                cell_str = str(cell).strip()
            if cell_str == po_str:
                result = {}
                for i in range(min(len(headers), len(row))):
                    if headers[i] is not None:
                        col_name = str(headers[i]).strip()
                        val = row[i]
                        if isinstance(val, datetime):
                            val = val.strftime("%Y-%m-%d") if not (val.hour or val.minute) \
                                else val.strftime("%Y-%m-%d %H:%M")
                        elif isinstance(val, date):
                            val = val.strftime("%Y-%m-%d")
                        result[col_name] = val
                return result, available_cols

        return None, available_cols
    finally:
        wb.close()


# =============================================================================
# Streamlit Page Config
# =============================================================================
st.set_page_config(
    page_title="Auto Form Generator - ปตท.",
    page_icon="📄",
    layout="centered",
    initial_sidebar_state="expanded",
)

# =============================================================================
# Session State
# =============================================================================
for key, default in [
    ("lookup_data", None),
    ("last_po", ""),
    ("status_msg", ""),
]:
    if key not in st.session_state:
        st.session_state[key] = default


# =============================================================================
# Sidebar: Excel data source
# =============================================================================
with st.sidebar:
    st.header("📊 แหล่งข้อมูล Excel")

    source = st.radio(
        "เลือกแหล่งข้อมูล:",
        ["📁 อัปโหลดไฟล์ Excel ของฉัน", "🎯 ใช้ไฟล์ตัวอย่าง (Demo)"],
        index=1,
    )

    excel_bytes = None
    sample_path = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                               'sample_po_data.xlsx')

    if source.startswith("📁"):
        uploaded = st.file_uploader("เลือกไฟล์ Excel (.xlsx)",
                                    type=['xlsx'],
                                    accept_multiple_files=False)
        if uploaded:
            excel_bytes = uploaded.read()
            st.success(f"✓ อัปโหลด: {uploaded.name}")
            st.caption(f"ขนาด: {len(excel_bytes) / 1024:.1f} KB")
    else:
        if os.path.exists(sample_path):
            with open(sample_path, 'rb') as f:
                excel_bytes = f.read()
            st.info("ใช้ไฟล์ตัวอย่าง")
            st.caption("ลอง PO: `4500987654`, `4500987655`, `4500987656`")
        else:
            st.warning("⚠ ไม่พบไฟล์ตัวอย่าง")

    st.divider()

    with st.expander("⚙️ ตั้งค่าขั้นสูง"):
        po_column = st.text_input("ชื่อคอลัมน์ PO ใน Excel:",
                                  value="poNo",
                                  help="ปกติ `poNo` — เปลี่ยนถ้า Excel ใช้ชื่ออื่น")
        sheet_name = st.text_input("ชื่อ Sheet:",
                                   value="",
                                   help="เว้นว่าง = ใช้ sheet แรก")

    st.divider()
    st.caption("📦 [GitHub Repo](https://github.com/Menixis/auto-form-generator)")
    st.caption("Made with ❤️ for ปตท.")


# =============================================================================
# Main UI
# =============================================================================
st.title("📄 ระบบสร้างเอกสารอัตโนมัติ")
st.caption("ปตท. — สร้างหนังสือราชการอัตโนมัติจากแบบฟอร์มมาตรฐาน 5 ประเภท "
           "พร้อมดึงข้อมูลจาก Excel")

# Form type
form_labels = [f["label"] for f in FORM_TYPES]
form_choice = st.selectbox("**ประเภทแบบฟอร์ม**", form_labels, index=0)
selected_form = next(f for f in FORM_TYPES if f["label"] == form_choice)

st.divider()

# PO Number + Fetch
col_po, col_btn = st.columns([3, 1])
with col_po:
    po_number = st.text_input(
        "**เลข PO**" + (" *" if selected_form["uses_po"] else ""),
        value=st.session_state.last_po,
        placeholder="เช่น 4500987654",
        key="po_input",
    )
with col_btn:
    st.markdown("<br>", unsafe_allow_html=True)  # spacer
    fetch_clicked = st.button("🔍 ดึงข้อมูล",
                              use_container_width=True,
                              type="secondary",
                              disabled=(excel_bytes is None))

# Fetch logic
if fetch_clicked:
    if not po_number.strip():
        st.warning("⚠ กรุณากรอกเลข PO ก่อน")
    elif not excel_bytes:
        st.error("ไม่มีไฟล์ Excel — กรุณาอัปโหลดไฟล์ที่ sidebar")
    else:
        with st.spinner("กำลังค้นหา..."):
            try:
                result, available_cols = lookup_po_in_bytes(
                    po_number.strip(),
                    excel_bytes,
                    sheet_name=sheet_name or None,
                    po_column=po_column.strip(),
                )
                if isinstance(result, dict) and "__error__" in result:
                    st.error(result["__error__"])
                    if available_cols:
                        st.info(f"คอลัมน์ที่พบในไฟล์: {', '.join(available_cols[:15])}"
                                + ("..." if len(available_cols) > 15 else ""))
                    st.session_state.lookup_data = None
                elif result:
                    st.session_state.lookup_data = result
                    st.session_state.last_po = po_number.strip()
                    vendor = result.get('VendorName', '(ไม่มีชื่อ vendor)')
                    st.success(f"✓ พบข้อมูล: **{vendor}**")
                else:
                    st.warning(f"⚠ ไม่พบ PO: {po_number}")
                    st.session_state.lookup_data = None
            except Exception as e:
                st.error(f"❌ ผิดพลาด: {e}")
                st.session_state.lookup_data = None

# Show fetched data preview
if st.session_state.lookup_data:
    with st.expander("📋 ข้อมูลที่ดึงจาก Excel", expanded=True):
        for col, label in DISPLAY_FIELDS:
            val = st.session_state.lookup_data.get(col)
            if val not in (None, ""):
                st.markdown(f"**{label}:** {val}")

st.divider()

# Other inputs
doc_number = st.text_input("**หมายเลขเอกสาร** *",
                           placeholder="เช่น PTT-001/2568")

if selected_form["uses_doc_type"]:
    doc_type = st.radio("**ประเภทเอกสาร**",
                        DOC_TYPE_OPTIONS,
                        horizontal=True)
else:
    doc_type = ""
    st.caption("ℹ️ แบบฟอร์มนี้ไม่ต้องระบุประเภทเอกสาร")

signer = st.text_input("**ผู้ลงนาม** *",
                       placeholder="ชื่อผู้มีอำนาจอนุมัติ หรือ ประธานกรรมการตรวจรับ")

st.divider()

# Generate button
generate_clicked = st.button("📄 สร้างเอกสาร",
                              type="primary",
                              use_container_width=True)

if generate_clicked:
    # Validate
    errors = []
    if not doc_number.strip():
        errors.append("กรุณากรอกหมายเลขเอกสาร")
    if not signer.strip():
        errors.append("กรุณากรอกชื่อผู้ลงนาม")
    if selected_form["uses_po"] and not po_number.strip():
        errors.append("กรุณากรอกเลข PO")

    if errors:
        for err in errors:
            st.error(f"⚠ {err}")
    else:
        template_path = os.path.join(
            os.path.dirname(os.path.abspath(__file__)),
            'templates',
            selected_form['filename']
        )

        if not os.path.exists(template_path):
            st.error(f"❌ ไม่พบไฟล์ template: {template_path}")
        else:
            try:
                with st.spinner("กำลังสร้างเอกสาร..."):
                    # ใช้ BytesIO เป็น output (python-docx รองรับ)
                    output_buffer = BytesIO()
                    log = generate_form(
                        form_id=selected_form['id'],
                        template_path=template_path,
                        output_path=output_buffer,
                        doc_number=doc_number.strip(),
                        po_number=po_number.strip(),
                        doc_type=doc_type if selected_form['uses_doc_type'] else "",
                        signer=signer.strip(),
                        lookup_data=st.session_state.lookup_data,
                    )
                    output_buffer.seek(0)
                    output_bytes = output_buffer.getvalue()

                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                safe_num = doc_number.replace('/', '-').replace('\\', '-').strip()
                filename = (f"form{selected_form['id']}_{safe_num}_{timestamp}.docx"
                            if safe_num else f"form{selected_form['id']}_{timestamp}.docx")

                st.success("✅ สร้างเอกสารสำเร็จ!")
                with st.expander("📋 รายละเอียดการแทนที่"):
                    for line in log:
                        st.markdown(f"- {line}")

                st.download_button(
                    label="⬇️ ดาวน์โหลดเอกสาร",
                    data=output_bytes,
                    file_name=filename,
                    mime="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
                    use_container_width=True,
                    type="primary",
                )
            except Exception as e:
                st.error(f"❌ ไม่สามารถสร้างเอกสารได้: {e}")

# Footer
st.divider()
st.caption("💡 **วิธีใช้งาน**: เลือกแบบฟอร์ม → กรอกเลข PO → กด \"ดึงข้อมูล\" → "
           "กรอกข้อมูลที่เหลือ → กด \"สร้างเอกสาร\" → ดาวน์โหลด")
